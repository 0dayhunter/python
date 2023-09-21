import openpyxl

# Load the Excel file
path = "abc.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

# Get the number of rows and columns
row, column = sheet.max_row, sheet.max_column
print("Total Rows:", row)
print("Total Columns:", column)

# Print the values of the first column
print("\nValue of first column")
for i in range(1, row + 1):
    print(sheet.cell(row=i, column=1).value)
z
# Print the values of the first row
print("\nValue of first row")
for i in range(1, column + 1):
    print(sheet.cell(row=2, column=i).value, end=" ")

# Create a new Excel workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Write values to cells in the new workbook
new_sheet.cell(row=1, column=1, value="Hello")
new_sheet.cell(row=1, column=2, value="World")
new_sheet['A2'] = "Welcome"
new_sheet['B2'] = "Everyone"

# Save the new workbook to a file
new_path = "new_abc.xlsx"
new_wb.save(new_path)
